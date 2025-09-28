"""
Advanced reporting services for generating comprehensive inventory reports
"""

import io
import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import streamlit as st

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel generation
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    PANDAS_AVAILABLE = True
    OPENPYXL_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    OPENPYXL_AVAILABLE = False

from core.db_utils import get_db_connection
from audit.services import get_audit_service


class ReportService:
    """Service for generating various inventory reports in PDF and Excel formats"""

    def __init__(self):
        self.audit_service = get_audit_service()

    def get_dependencies_status(self) -> Dict[str, bool]:
        """Check if required dependencies are available"""
        return {
            "reportlab": REPORTLAB_AVAILABLE,
            "pandas": PANDAS_AVAILABLE,
            "openpyxl": OPENPYXL_AVAILABLE
        }

    def get_missing_dependencies(self) -> List[str]:
        """Get list of missing dependencies"""
        status = self.get_dependencies_status()
        missing = []
        if not status["reportlab"]:
            missing.append("reportlab")
        if not status["pandas"]:
            missing.append("pandas")
        if not status["openpyxl"]:
            missing.append("openpyxl")
        return missing

    def get_summary_data(self) -> Dict[str, Any]:
        """Get summary data for reports"""
        try:
            with get_db_connection() as conn:
                # Hardware summary
                hardware_summary = conn.execute("""
                    SELECT
                        kategorie,
                        COUNT(*) as anzahl,
                        SUM(preis) as gesamtwert,
                        AVG(preis) as durchschnittspreis
                    FROM hardware_inventar
                    GROUP BY kategorie
                """).fetchall()

                # Cable summary
                cable_summary = conn.execute("""
                    SELECT
                        typ,
                        COUNT(*) as anzahl,
                        SUM(laenge) as gesamtlaenge,
                        AVG(laenge) as durchschnittslaenge
                    FROM kabel_inventar
                    GROUP BY typ
                """).fetchall()

                # Location summary
                location_summary = conn.execute("""
                    SELECT
                        l.name as standort,
                        COUNT(DISTINCT h.id) + COUNT(DISTINCT k.id) as gesamtitems,
                        COUNT(DISTINCT h.id) as hardware_anzahl,
                        COUNT(DISTINCT k.id) as kabel_anzahl
                    FROM standorte l
                    LEFT JOIN hardware_inventar h ON l.id = h.standort_id
                    LEFT JOIN kabel_inventar k ON l.id = k.standort_id
                    GROUP BY l.id, l.name
                """).fetchall()

                # Status distribution
                status_distribution = conn.execute("""
                    SELECT status, COUNT(*) as anzahl
                    FROM (
                        SELECT status FROM hardware_inventar
                        UNION ALL
                        SELECT status FROM kabel_inventar
                    ) combined
                    GROUP BY status
                """).fetchall()

                # Total values
                total_hardware_value = conn.execute("""
                    SELECT COALESCE(SUM(preis), 0) as total FROM hardware_inventar
                """).fetchone()[0]

                total_items = conn.execute("""
                    SELECT
                        (SELECT COUNT(*) FROM hardware_inventar) +
                        (SELECT COUNT(*) FROM kabel_inventar) as total
                """).fetchone()[0]

                return {
                    "hardware_summary": [dict(row) for row in hardware_summary],
                    "cable_summary": [dict(row) for row in cable_summary],
                    "location_summary": [dict(row) for row in location_summary],
                    "status_distribution": [dict(row) for row in status_distribution],
                    "total_hardware_value": total_hardware_value,
                    "total_items": total_items,
                    "generated_at": datetime.now()
                }
        except Exception as e:
            st.error(f"Fehler beim Abrufen der Zusammenfassungsdaten: {e}")
            return {}

    def get_detailed_inventory_data(self, item_type: str = "all") -> Dict[str, Any]:
        """Get detailed inventory data"""
        try:
            with get_db_connection() as conn:
                data = {"generated_at": datetime.now()}

                if item_type in ["all", "hardware"]:
                    hardware_data = conn.execute("""
                        SELECT
                            h.*,
                            s.name as standort_name,
                            s.beschreibung as standort_beschreibung
                        FROM hardware_inventar h
                        LEFT JOIN standorte s ON h.standort_id = s.id
                        ORDER BY h.kategorie, h.modell
                    """).fetchall()
                    data["hardware"] = [dict(row) for row in hardware_data]

                if item_type in ["all", "cables"]:
                    cable_data = conn.execute("""
                        SELECT
                            k.*,
                            s.name as standort_name,
                            s.beschreibung as standort_beschreibung
                        FROM kabel_inventar k
                        LEFT JOIN standorte s ON k.standort_id = s.id
                        ORDER BY k.typ, k.kategorie
                    """).fetchall()
                    data["cables"] = [dict(row) for row in cable_data]

                return data
        except Exception as e:
            st.error(f"Fehler beim Abrufen der detaillierten Inventardaten: {e}")
            return {}

    def get_valuation_data(self) -> Dict[str, Any]:
        """Get valuation and financial data"""
        try:
            with get_db_connection() as conn:
                # Category valuations
                category_valuations = conn.execute("""
                    SELECT
                        kategorie,
                        COUNT(*) as anzahl,
                        SUM(preis) as gesamtwert,
                        AVG(preis) as durchschnittspreis,
                        MIN(preis) as minpreis,
                        MAX(preis) as maxpreis
                    FROM hardware_inventar
                    WHERE preis IS NOT NULL AND preis > 0
                    GROUP BY kategorie
                    ORDER BY gesamtwert DESC
                """).fetchall()

                # Location valuations
                location_valuations = conn.execute("""
                    SELECT
                        s.name as standort,
                        COUNT(h.id) as hardware_anzahl,
                        SUM(h.preis) as gesamtwert,
                        AVG(h.preis) as durchschnittspreis
                    FROM standorte s
                    LEFT JOIN hardware_inventar h ON s.id = h.standort_id AND h.preis > 0
                    GROUP BY s.id, s.name
                    HAVING COUNT(h.id) > 0
                    ORDER BY gesamtwert DESC
                """).fetchall()

                # Age-based valuations (depreciation analysis)
                age_valuations = conn.execute("""
                    SELECT
                        CASE
                            WHEN julianday('now') - julianday(anschaffungsdatum) < 365 THEN 'Unter 1 Jahr'
                            WHEN julianday('now') - julianday(anschaffungsdatum) < 1095 THEN '1-3 Jahre'
                            WHEN julianday('now') - julianday(anschaffungsdatum) < 1825 THEN '3-5 Jahre'
                            ELSE 'Über 5 Jahre'
                        END as altersgruppe,
                        COUNT(*) as anzahl,
                        SUM(preis) as gesamtwert,
                        AVG(preis) as durchschnittspreis
                    FROM hardware_inventar
                    WHERE anschaffungsdatum IS NOT NULL AND preis > 0
                    GROUP BY altersgruppe
                    ORDER BY
                        CASE altersgruppe
                            WHEN 'Unter 1 Jahr' THEN 1
                            WHEN '1-3 Jahre' THEN 2
                            WHEN '3-5 Jahre' THEN 3
                            ELSE 4
                        END
                """).fetchall()

                return {
                    "category_valuations": [dict(row) for row in category_valuations],
                    "location_valuations": [dict(row) for row in location_valuations],
                    "age_valuations": [dict(row) for row in age_valuations],
                    "generated_at": datetime.now()
                }
        except Exception as e:
            st.error(f"Fehler beim Abrufen der Bewertungsdaten: {e}")
            return {}

    def get_maintenance_data(self) -> Dict[str, Any]:
        """Get maintenance and warranty data"""
        try:
            with get_db_connection() as conn:
                # Warranty status
                warranty_status = conn.execute("""
                    SELECT
                        CASE
                            WHEN garantie_ende IS NULL THEN 'Keine Garantie'
                            WHEN date(garantie_ende) < date('now') THEN 'Abgelaufen'
                            WHEN date(garantie_ende) < date('now', '+90 days') THEN 'Läuft bald ab'
                            ELSE 'Aktiv'
                        END as garantie_status,
                        COUNT(*) as anzahl,
                        SUM(preis) as gesamtwert
                    FROM hardware_inventar
                    GROUP BY garantie_status
                """).fetchall()

                # Upcoming warranty expirations
                warranty_expiring = conn.execute("""
                    SELECT
                        seriennummer,
                        hersteller,
                        modell,
                        garantie_ende,
                        preis,
                        julianday(garantie_ende) - julianday('now') as tage_bis_ablauf
                    FROM hardware_inventar
                    WHERE garantie_ende IS NOT NULL
                    AND date(garantie_ende) >= date('now')
                    AND date(garantie_ende) <= date('now', '+180 days')
                    ORDER BY garantie_ende
                """).fetchall()

                # Age distribution
                age_distribution = conn.execute("""
                    SELECT
                        CASE
                            WHEN anschaffungsdatum IS NULL THEN 'Unbekannt'
                            WHEN julianday('now') - julianday(anschaffungsdatum) < 365 THEN 'Unter 1 Jahr'
                            WHEN julianday('now') - julianday(anschaffungsdatum) < 1095 THEN '1-3 Jahre'
                            WHEN julianday('now') - julianday(anschaffungsdatum) < 1825 THEN '3-5 Jahre'
                            ELSE 'Über 5 Jahre'
                        END as alter,
                        COUNT(*) as anzahl
                    FROM hardware_inventar
                    GROUP BY alter
                    ORDER BY
                        CASE alter
                            WHEN 'Unter 1 Jahr' THEN 1
                            WHEN '1-3 Jahre' THEN 2
                            WHEN '3-5 Jahre' THEN 3
                            WHEN 'Über 5 Jahre' THEN 4
                            ELSE 5
                        END
                """).fetchall()

                return {
                    "warranty_status": [dict(row) for row in warranty_status],
                    "warranty_expiring": [dict(row) for row in warranty_expiring],
                    "age_distribution": [dict(row) for row in age_distribution],
                    "generated_at": datetime.now()
                }
        except Exception as e:
            st.error(f"Fehler beim Abrufen der Wartungsdaten: {e}")
            return {}

    def generate_summary_report_pdf(self, data: Dict[str, Any]) -> io.BytesIO:
        """Generate summary report in PDF format"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab ist nicht installiert")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph("Inventory Management System - Zusammenfassungsbericht", title_style))
        story.append(Spacer(1, 20))

        # Generation info
        info_style = styles['Normal']
        story.append(Paragraph(f"Generiert am: {data['generated_at'].strftime('%d.%m.%Y %H:%M')}", info_style))
        story.append(Spacer(1, 20))

        # Key metrics
        metrics_data = [
            ['Metrik', 'Wert'],
            ['Gesamtanzahl Items', str(data.get('total_items', 0))],
            ['Gesamtwert Hardware', f"€{data.get('total_hardware_value', 0):,.2f}"],
            ['Anzahl Kategorien', str(len(data.get('hardware_summary', [])))],
            ['Anzahl Standorte', str(len(data.get('location_summary', [])))]
        ]

        metrics_table = Table(metrics_data)
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 30))

        # Hardware summary
        if data.get('hardware_summary'):
            story.append(Paragraph("Hardware Zusammenfassung nach Kategorie", styles['Heading2']))
            hardware_data = [['Kategorie', 'Anzahl', 'Gesamtwert', 'Durchschnittspreis']]
            for item in data['hardware_summary']:
                hardware_data.append([
                    item['kategorie'],
                    str(item['anzahl']),
                    f"€{item['gesamtwert']:,.2f}" if item['gesamtwert'] else "€0.00",
                    f"€{item['durchschnittspreis']:,.2f}" if item['durchschnittspreis'] else "€0.00"
                ])

            hardware_table = Table(hardware_data)
            hardware_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(hardware_table)
            story.append(Spacer(1, 20))

        # Status distribution
        if data.get('status_distribution'):
            story.append(Paragraph("Status Verteilung", styles['Heading2']))
            status_data = [['Status', 'Anzahl']]
            for item in data['status_distribution']:
                status_data.append([item['status'], str(item['anzahl'])])

            status_table = Table(status_data)
            status_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(status_table)

        doc.build(story)
        buffer.seek(0)
        return buffer

    def generate_detailed_report_excel(self, data: Dict[str, Any]) -> io.BytesIO:
        """Generate detailed inventory report in Excel format"""
        if not PANDAS_AVAILABLE or not OPENPYXL_AVAILABLE:
            raise ImportError("Pandas oder openpyxl ist nicht installiert")

        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metrik': ['Generiert am', 'Gesamtanzahl Items', 'Hardware Items', 'Kabel Items'],
                'Wert': [
                    data['generated_at'].strftime('%d.%m.%Y %H:%M'),
                    len(data.get('hardware', [])) + len(data.get('cables', [])),
                    len(data.get('hardware', [])),
                    len(data.get('cables', []))
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Zusammenfassung', index=False)

            # Hardware sheet
            if data.get('hardware'):
                hardware_df = pd.DataFrame(data['hardware'])
                # Select and rename columns for better readability
                hardware_columns = {
                    'seriennummer': 'Seriennummer',
                    'hersteller': 'Hersteller',
                    'modell': 'Modell',
                    'kategorie': 'Kategorie',
                    'preis': 'Preis',
                    'anschaffungsdatum': 'Anschaffungsdatum',
                    'garantie_ende': 'Garantie Ende',
                    'status': 'Status',
                    'standort_name': 'Standort'
                }
                hardware_display_df = hardware_df[[col for col in hardware_columns.keys() if col in hardware_df.columns]]
                hardware_display_df = hardware_display_df.rename(columns=hardware_columns)
                hardware_display_df.to_excel(writer, sheet_name='Hardware Inventar', index=False)

            # Cables sheet
            if data.get('cables'):
                cables_df = pd.DataFrame(data['cables'])
                cable_columns = {
                    'seriennummer': 'Seriennummer',
                    'typ': 'Typ',
                    'kategorie': 'Kategorie',
                    'laenge': 'Länge (m)',
                    'farbe': 'Farbe',
                    'anschaffungsdatum': 'Anschaffungsdatum',
                    'status': 'Status',
                    'standort_name': 'Standort'
                }
                cables_display_df = cables_df[[col for col in cable_columns.keys() if col in cables_df.columns]]
                cables_display_df = cables_display_df.rename(columns=cable_columns)
                cables_display_df.to_excel(writer, sheet_name='Kabel Inventar', index=False)

            # Format the Excel file
            workbook = writer.book
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Style header row
                if worksheet.max_row > 0:
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)

        buffer.seek(0)
        return buffer

    def generate_valuation_report_pdf(self, data: Dict[str, Any]) -> io.BytesIO:
        """Generate valuation report in PDF format"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab ist nicht installiert")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Bewertungsbericht - Asset Valuation", title_style))
        story.append(Spacer(1, 20))

        # Generation info
        story.append(Paragraph(f"Generiert am: {data['generated_at'].strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))

        # Category valuations
        if data.get('category_valuations'):
            story.append(Paragraph("Bewertung nach Kategorien", styles['Heading2']))
            cat_data = [['Kategorie', 'Anzahl', 'Gesamtwert', 'Durchschnitt', 'Min', 'Max']]
            total_value = 0
            for item in data['category_valuations']:
                cat_data.append([
                    item['kategorie'],
                    str(item['anzahl']),
                    f"€{item['gesamtwert']:,.2f}",
                    f"€{item['durchschnittspreis']:,.2f}",
                    f"€{item['minpreis']:,.2f}",
                    f"€{item['maxpreis']:,.2f}"
                ])
                total_value += item['gesamtwert']

            # Add total row
            cat_data.append(['GESAMT', '', f"€{total_value:,.2f}", '', '', ''])

            cat_table = Table(cat_data)
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(cat_table)
            story.append(Spacer(1, 20))

        # Age-based valuations
        if data.get('age_valuations'):
            story.append(Paragraph("Bewertung nach Alter (Abschreibung)", styles['Heading2']))
            age_data = [['Altersgruppe', 'Anzahl', 'Gesamtwert', 'Durchschnitt']]
            for item in data['age_valuations']:
                age_data.append([
                    item['altersgruppe'],
                    str(item['anzahl']),
                    f"€{item['gesamtwert']:,.2f}",
                    f"€{item['durchschnittspreis']:,.2f}"
                ])

            age_table = Table(age_data)
            age_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(age_table)

        doc.build(story)
        buffer.seek(0)
        return buffer


# Global service instance
_report_service = None


def get_report_service() -> ReportService:
    """Get global report service instance"""
    global _report_service
    if _report_service is None:
        _report_service = ReportService()
    return _report_service
